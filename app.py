"""
📦 調貨建議生成系統 v1.72
零售庫存調貨建議生成系統

開發者: MiniMax Agent
創建日期: 2025-09-18
更新日期: 2025-09-30
v1.72: 修正模式B同店舖同SKU轉出/接收衝突問題
"""

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import warnings
warnings.filterwarnings('ignore')

# 設置頁面配置
st.set_page_config(
    page_title="調貨建議生成系統",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 設置matplotlib中文字體
plt.switch_backend("Agg")
plt.rcParams["font.sans-serif"] = ["Noto Sans CJK SC", "WenQuanYi Zen Hei", "PingFang SC", "Arial Unicode MS", "Hiragino Sans GB"]
plt.rcParams["axes.unicode_minus"] = False

# 全局樣式設定
st.markdown("""
<style>
.main-title {
    font-size: 2.5rem;
    color: #1f77b4;
    text-align: center;
    margin-bottom: 2rem;
}
.section-header {
    background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
    color: white;
    padding: 10px;
    border-radius: 5px;
    margin: 10px 0px;
}
.metric-card {
    background: #f0f2f6;
    padding: 15px;
    border-radius: 10px;
    border-left: 4px solid #1f77b4;
}
</style>
""", unsafe_allow_html=True)

class TransferRecommendationSystem:
    """調貨建議系統核心類"""
    
    def __init__(self):
        self.df = None
        self.transfer_suggestions = None
        self.statistics = None
        self.mode = "A"  # A: 保守轉貨, B: 加強轉貨
        
    def calculate_preliminary_statistics(self):
        """計算預先統計數據（預計需求、轉出、接收數量）"""
        if self.df is None:
            return {}
        
        # 計算A模式和B模式的預計數量
        stats_a = self._calculate_mode_statistics("A")
        stats_b = self._calculate_mode_statistics("B")
        
        return {
            'conservative': stats_a,
            'enhanced': stats_b
        }
    
    def _calculate_mode_statistics(self, mode):
        """計算指定模式的統計數據"""
        transfer_candidates = self.identify_transfer_candidates(mode)
        receive_candidates = self.identify_receive_candidates()
        
        total_transfer = sum(candidate['Transfer_Qty'] for candidate in transfer_candidates)
        total_receive = sum(candidate['Need_Qty'] for candidate in receive_candidates)
        
        return {
            'estimated_transfer': total_transfer,
            'estimated_receive': total_receive,
            'estimated_demand': total_receive  # 需求等於接收
        }
    
    def load_and_preprocess_data(self, uploaded_file):
        """載入和預處理數據"""
        try:
            # 讀取Excel文件 - 添加編碼處理
            df = pd.read_excel(uploaded_file)
            
            # 清理列名中的異常字符
            df.columns = df.columns.astype(str)
            cleaned_columns = []
            for col in df.columns:
                # 更严格的字符清理：只保留字母、数字、中文、常用符号
                import re
                # 首先移除明显的异常字符模式
                cleaned_col = str(col)
                
                # 检测并移除类似 "key匡省得斗儿俩焯v_right" 这样的异常模式
                if re.search(r'key.*v_right|[\u0000-\u001f\u007f-\u009f]', cleaned_col):
                    # 如果包含明显异常模式，尝试提取有意义的部分或使用默认名称
                    cleaned_col = f"Unknown_Column_{len(cleaned_columns)}"
                else:
                    # 正常清理：只保留安全字符
                    cleaned_col = re.sub(r'[^\w\s\u4e00-\u9fff\-_.()]', '', cleaned_col)
                    cleaned_col = cleaned_col.strip()
                    if not cleaned_col:
                        cleaned_col = f"Column_{len(cleaned_columns)}"
                
                cleaned_columns.append(cleaned_col)
            
            df.columns = cleaned_columns
            
            # 驗證必需欄位
            required_columns = [
                'Article', 'Article Description', 'RP Type', 'Site', 'OM', 
                'MOQ', 'SaSa Net Stock', 'Pending Received', 'Safety Stock', 
                'Last Month Sold Qty', 'MTD Sold Qty'
            ]
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f"缺少必需欄位: {', '.join(missing_columns)}")
            
            # 數據預處理
            df['Article'] = df['Article'].astype(str)
            
            # 處理數值欄位
            numeric_columns = ['MOQ', 'SaSa Net Stock', 'Pending Received', 'Safety Stock', 
                             'Last Month Sold Qty', 'MTD Sold Qty']
            
            df['Notes'] = ""  # 添加備註欄位
            
            for col in numeric_columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
                
                # 負值修正
                mask_negative = df[col] < 0
                if mask_negative.any():
                    df.loc[mask_negative, 'Notes'] += f"{col}負值修正為0; "
                    df.loc[mask_negative, col] = 0
                
                # 銷量異常值處理
                if 'Sold Qty' in col:
                    mask_extreme = df[col] > 100000
                    if mask_extreme.any():
                        df.loc[mask_extreme, 'Notes'] += f"{col}異常值>100000修正; "
                        df.loc[mask_extreme, col] = 100000
            
            # 字串欄位處理 - 添加異常字符清理
            string_columns = ['Article Description', 'RP Type', 'Site', 'OM']
            for col in string_columns:
                if col in df.columns:  # 確保列存在
                    df[col] = df[col].fillna('').astype(str)
                    # 更严格的數據內容清理
                    import re
                    df[col] = df[col].apply(lambda x: 
                        re.sub(r'[^\w\s\u4e00-\u9fff\-_()./]', '', str(x)).strip() 
                        if not re.search(r'key.*v_right|[\u0000-\u001f\u007f-\u009f]', str(x)) 
                        else 'CLEANED_DATA'
                    )
                else:
                    df[col] = ''  # 如果列不存在，創建空列
            
            # 驗證RP Type值
            valid_rp_types = ['ND', 'RF']
            invalid_rp_mask = ~df['RP Type'].isin(valid_rp_types)
            if invalid_rp_mask.any():
                df.loc[invalid_rp_mask, 'Notes'] += "RP Type無效值; "
                df.loc[invalid_rp_mask, 'RP Type'] = 'RF'  # 預設為RF
            
            self.df = df
            
            # 計算預先統計
            self.preliminary_stats = self.calculate_preliminary_statistics()
            
            return True, f"成功載入 {len(df)} 筆記錄"
            
        except Exception as e:
            return False, f"數據載入失敗: {str(e)}"
    
    def calculate_effective_sales(self, row):
        """計算有效銷量"""
        if row['Last Month Sold Qty'] > 0:
            return row['Last Month Sold Qty']
        else:
            return row['MTD Sold Qty']
    
    def identify_transfer_candidates(self, mode="A"):
        """識別轉出候選"""
        candidates = []
        
        for _, row in self.df.iterrows():
            article = row['Article']
            current_stock = row['SaSa Net Stock']
            pending = row['Pending Received']
            safety_stock = row['Safety Stock']
            rp_type = row['RP Type']
            effective_sales = self.calculate_effective_sales(row)
            moq = row['MOQ']
            
            # 計算該產品的最高銷量
            product_data = self.df[self.df['Article'] == article]
            max_sales = product_data.apply(self.calculate_effective_sales, axis=1).max()
            
            # ND類型完全轉出 (優先順序1)
            if rp_type == 'ND' and current_stock > 0:
                # ND類型完全轉出，剩餘庫存為0
                remaining_stock = 0
                
                candidates.append({
                    'Article': article,
                    'Site': row['Site'],
                    'OM': row['OM'],
                    'Transfer_Qty': current_stock,
                    'Type': 'ND轉出',
                    'Priority': 1,
                    'Original_Stock': current_stock,
                    'Safety_Stock': safety_stock,
                    'MOQ': moq,
                    'Effective_Sales': effective_sales,
                    'Total_Available': current_stock + pending,
                    'Remaining_Stock': remaining_stock  # 添加剩餘庫存信息
                })
            
            # RF類型轉出 (優先順序2)
            elif rp_type == 'RF' and effective_sales < max_sales:
                total_available = current_stock + pending
                
                if mode == "A":  # 保守轉貨
                    if total_available > safety_stock:
                        base_transfer = total_available - safety_stock
                        limit_transfer = max(int(total_available * 0.2), 2)
                        actual_transfer = min(base_transfer, limit_transfer)
                        actual_transfer = min(actual_transfer, current_stock)
                        
                        if actual_transfer > 0:
                            # 計算轉出後剩餘庫存
                            remaining_stock = current_stock - actual_transfer
                            
                            candidates.append({
                                'Article': article,
                                'Site': row['Site'],
                                'OM': row['OM'],
                                'Transfer_Qty': actual_transfer,
                                'Type': 'RF過剩轉出',
                                'Priority': 2,
                                'Original_Stock': current_stock,
                                'Safety_Stock': safety_stock,
                                'MOQ': moq,
                                'Effective_Sales': effective_sales,
                                'Total_Available': total_available,
                                'Remaining_Stock': remaining_stock  # 添加剩餘庫存信息
                            })
                            
                elif mode == "B":  # 加強轉貨
                    moq_threshold = moq + 1
                    if total_available > moq_threshold:
                        base_transfer = total_available - moq_threshold
                        limit_transfer = max(int(total_available * 0.5), 2)
                        actual_transfer = min(base_transfer, limit_transfer)
                        actual_transfer = min(actual_transfer, current_stock)
                        
                        if actual_transfer > 0:
                            # 計算轉出後剩餘庫存
                            remaining_stock = current_stock - actual_transfer
                            
                            # 根據剩餘庫存與Safety stock關係確定轉出類型
                            if remaining_stock >= safety_stock:
                                transfer_type = 'RF過剩轉出'  # 剩餘庫存不會低於Safety stock
                            else:
                                transfer_type = 'RF加強轉出'  # 剩餘庫存會低於Safety stock
                            
                            candidates.append({
                                'Article': article,
                                'Site': row['Site'],
                                'OM': row['OM'],
                                'Transfer_Qty': actual_transfer,
                                'Type': transfer_type,
                                'Priority': 2,
                                'Original_Stock': current_stock,
                                'Safety_Stock': safety_stock,
                                'MOQ': moq,
                                'Effective_Sales': effective_sales,
                                'Total_Available': total_available,
                                'Remaining_Stock': remaining_stock  # 添加剩餘庫存信息
                            })
        
        # 按有效銷量排序（低銷量優先轉出）
        candidates.sort(key=lambda x: (x['Priority'], x['Effective_Sales']))
        return candidates
    
    def identify_receive_candidates(self):
        """識別接收候選 - v1.71 優化：添加SasaNet調撥接收條件"""
        candidates = []
        
        for _, row in self.df.iterrows():
            article = row['Article']
            current_stock = row['SaSa Net Stock']
            pending = row['Pending Received']
            safety_stock = row['Safety Stock']
            rp_type = row['RP Type']
            effective_sales = self.calculate_effective_sales(row)
            site = row['Site']
            
            if rp_type == 'RF':
                # 計算該產品的最高銷量
                product_data = self.df[self.df['Article'] == article]
                max_sales = product_data.apply(self.calculate_effective_sales, axis=1).max()
                
                # 緊急缺貨補貨 (優先順序1)
                if current_stock == 0 and effective_sales > 0:
                    candidates.append({
                        'Article': article,
                        'Site': site,
                        'OM': row['OM'],
                        'Need_Qty': safety_stock,
                        'Type': '緊急缺貨補貨',
                        'Priority': 1,
                        'Current_Stock': current_stock,
                        'Safety_Stock': safety_stock,
                        'Effective_Sales': effective_sales,
                        'Pending_Received': pending,
                        'Total_Available': current_stock + pending
                    })
                
                # v1.71 新增：SasaNet 調撥接收條件 (優先順序2)
                elif (current_stock + pending) < safety_stock and current_stock > 0:
                    need_qty = safety_stock - (current_stock + pending)
                    if need_qty > 0:
                        candidates.append({
                            'Article': article,
                            'Site': site,
                            'OM': row['OM'],
                            'Need_Qty': need_qty,
                            'Type': 'SasaNet調撥接收',
                            'Priority': 2,
                            'Current_Stock': current_stock,
                            'Safety_Stock': safety_stock,
                            'Effective_Sales': effective_sales,
                            'Pending_Received': pending,
                            'Total_Available': current_stock + pending
                        })
                
                # 潛在缺貨補貨 (優先順序3)
                elif (current_stock + pending) < safety_stock and effective_sales == max_sales:
                    need_qty = safety_stock - (current_stock + pending)
                    if need_qty > 0:
                        candidates.append({
                            'Article': article,
                            'Site': site,
                            'OM': row['OM'],
                            'Need_Qty': need_qty,
                            'Type': '潛在缺貨補貨',
                            'Priority': 3,
                            'Current_Stock': current_stock,
                            'Safety_Stock': safety_stock,
                            'Effective_Sales': effective_sales,
                            'Pending_Received': pending,
                            'Total_Available': current_stock + pending
                        })
        
        # 按優先順序和銷量排序
        candidates.sort(key=lambda x: (x['Priority'], -x['Effective_Sales']))
        return candidates
    
    def resolve_same_store_conflicts(self, transfer_candidates, receive_candidates):
        """
        解決同店舖同SKU衝突問題 - v1.72
        當同一店舖的同一SKU既被識別為轉出又被識別為接收時，
        優先保持接收需求，移除轉出候選
        """
        # 創建接收候選的查找表 (Store, Article, OM) -> 接收信息
        receive_lookup = {}
        for receive in receive_candidates:
            key = (receive['Site'], receive['Article'], receive['OM'])
            receive_lookup[key] = receive
        
        # 檢查轉出候選中的衝突並移除
        filtered_transfer_candidates = []
        conflicts_resolved = []
        
        for transfer in transfer_candidates:
            key = (transfer['Site'], transfer['Article'], transfer['OM'])
            
            if key in receive_lookup:
                # 發現衝突：同店舖同SKU既要轉出又要接收
                receive_info = receive_lookup[key]
                conflicts_resolved.append({
                    'site': transfer['Site'],
                    'article': transfer['Article'],
                    'om': transfer['OM'],
                    'transfer_qty': transfer['Transfer_Qty'],
                    'transfer_type': transfer['Type'],
                    'receive_qty': receive_info['Need_Qty'],
                    'receive_type': receive_info['Type']
                })
                # 不添加到過濾後的轉出候選中（優先保持接收）
                continue
            else:
                # 無衝突，保留轉出候選
                filtered_transfer_candidates.append(transfer)
        
        # 記錄衝突解決信息
        if conflicts_resolved:
            print(f"🔧 解決了 {len(conflicts_resolved)} 個同店舖同SKU衝突：")
            for conflict in conflicts_resolved:
                print(f"   {conflict['site']} - {conflict['article']}: 移除轉出{conflict['transfer_qty']}件({conflict['transfer_type']})，保持接收{conflict['receive_qty']}件({conflict['receive_type']})")
        
        return filtered_transfer_candidates, receive_candidates
    
    def match_transfer_suggestions(self, transfer_candidates, receive_candidates):
        """匹配調貨建議 - 優化同店舖RF轉出"""
        suggestions = []
        
        # 創建可變的候選列表副本
        available_transfers = transfer_candidates.copy()
        available_receives = receive_candidates.copy()
        
        # 先處理ND轉出（優先級最高）
        self._match_nd_transfers(available_transfers, available_receives, suggestions)
        
        # 再處理RF轉出，優化同店舖轉出
        self._match_rf_transfers_optimized(available_transfers, available_receives, suggestions)
        
        return suggestions
    
    def _match_nd_transfers(self, available_transfers, available_receives, suggestions):
        """處理ND轉出（最高優先級）"""
        for receive in available_receives[:]:
            if receive['Need_Qty'] <= 0:
                continue
                
            for i, transfer in enumerate(available_transfers):
                if (transfer['Type'] == 'ND轉出' and
                    transfer['Article'] == receive['Article'] and 
                    transfer['OM'] == receive['OM'] and 
                    transfer['Site'] != receive['Site'] and
                    transfer['Transfer_Qty'] > 0):
                    
                    actual_qty = min(transfer['Transfer_Qty'], receive['Need_Qty'])
                    
                    if actual_qty > 0:
                        suggestions.append(self._create_suggestion(transfer, receive, actual_qty))
                        available_transfers[i]['Transfer_Qty'] -= actual_qty
                        receive['Need_Qty'] -= actual_qty
                        
                        if receive['Need_Qty'] <= 0:
                            break
    
    def _match_rf_transfers_optimized(self, available_transfers, available_receives, suggestions):
        """處理RF轉出 - 優化存貨優先、同店舖轉出、避免單件"""
        # 將RF轉出按店舖分組
        rf_transfers_by_site = {}
        for i, transfer in enumerate(available_transfers):
            if transfer['Type'] in ['RF過剩轉出', 'RF加強轉出'] and transfer['Transfer_Qty'] > 0:
                site = transfer['Site']
                if site not in rf_transfers_by_site:
                    rf_transfers_by_site[site] = []
                rf_transfers_by_site[site].append((i, transfer))
        
        # 計算每個店舖的綜合優先級
        site_priority = []
        for site, transfers in rf_transfers_by_site.items():
            # 統計活躍轉出項目
            active_transfers = [t for i, t in transfers if t['Transfer_Qty'] > 0]
            
            if not active_transfers:
                continue
                
            # 計算優先級指標
            active_items = len(active_transfers)  # 可轉出品項數
            total_stock = sum(t['Original_Stock'] for t in active_transfers)  # 總存貨
            total_qty = sum(t['Transfer_Qty'] for t in active_transfers)  # 總可轉出數量
            multi_piece_items = len([t for t in active_transfers if t['Transfer_Qty'] >= 2])  # 可2件以上轉出的品項數
            
            # 綜合優先級：(可2件品項數, 總存貨, 品項數, 總轉出數量)
            priority = (multi_piece_items, total_stock, active_items, total_qty)
            site_priority.append((site, priority, transfers))
        
        # 按綜合優先級排序
        # 1. 可2件以上轉出品項數多的優先
        # 2. 總存貨多的優先  
        # 3. 品項數多的優先
        # 4. 總轉出數量多的優先
        site_priority.sort(key=lambda x: x[1], reverse=True)
        
        # 按優先順序處理每個店舖的轉出
        for site, priority_metrics, transfers in site_priority:
            # 在店舖內按存貨量排序轉出項目
            transfers_sorted = []
            for i, transfer in transfers:
                if transfer['Transfer_Qty'] > 0:
                    # 優先級：(可轉出數量>=2, 原始存貨, 轉出數量)
                    can_multi = 1 if transfer['Transfer_Qty'] >= 2 else 0
                    item_priority = (can_multi, transfer['Original_Stock'], transfer['Transfer_Qty'])
                    transfers_sorted.append((item_priority, i, transfer))
            
            # 按項目優先級排序
            transfers_sorted.sort(key=lambda x: x[0], reverse=True)
            
            # 處理該店舖的所有轉出需求
            for receive in available_receives[:]:
                if receive['Need_Qty'] <= 0:
                    continue
                    
                # 按優先級順序查找匹配的轉出項目
                for item_priority, i, transfer in transfers_sorted:
                    if (transfer['Article'] == receive['Article'] and 
                        transfer['OM'] == receive['OM'] and 
                        transfer['Site'] != receive['Site'] and
                        transfer['Transfer_Qty'] > 0):
                        
                        actual_qty = min(transfer['Transfer_Qty'], receive['Need_Qty'])
                        
                        # 智能數量優化
                        if actual_qty == 1:
                            # 如果只有1件且該轉出項目有足夠庫存，嘗試調高到2件
                            if transfer['Transfer_Qty'] >= 2:
                                after_transfer_stock = transfer['Original_Stock'] - 2
                                if after_transfer_stock >= transfer['Safety_Stock']:
                                    actual_qty = 2
                            else:
                                # 如果真的只能轉1件，檢查是否有其他店舖可以轉2件以上
                                if self._has_better_multi_piece_option(receive, rf_transfers_by_site, site):
                                    continue  # 跳過此次1件轉出，等待更好的選項
                        
                        if actual_qty > 0:
                            suggestions.append(self._create_suggestion(transfer, receive, actual_qty))
                            available_transfers[i]['Transfer_Qty'] -= actual_qty
                            transfer['Transfer_Qty'] -= actual_qty  # 同步更新本地副本
                            receive['Need_Qty'] -= actual_qty
                            
                            if receive['Need_Qty'] <= 0:
                                break
                                
    def _has_better_multi_piece_option(self, receive, rf_transfers_by_site, current_site):
        """檢查是否有其他店舖能提供2件以上的轉出選項"""
        for site, transfers in rf_transfers_by_site.items():
            if site == current_site:
                continue
                
            for i, transfer in transfers:
                if (transfer['Article'] == receive['Article'] and 
                    transfer['OM'] == receive['OM'] and 
                    transfer['Site'] != receive['Site'] and
                    transfer['Transfer_Qty'] >= 2):
                    return True
        return False
    
    def _create_suggestion(self, transfer, receive, actual_qty):
        """創建調貨建議記錄"""
        return {
            'Article': transfer['Article'],
            'OM': transfer['OM'],
            'Transfer_Site': transfer['Site'],
            'Receive_Site': receive['Site'],
            'Transfer_Qty': actual_qty,
            'Transfer_Type': transfer['Type'],
            'Receive_Type': receive['Type'],
            'Original_Stock': transfer['Original_Stock'],
            'After_Transfer_Stock': transfer['Original_Stock'] - actual_qty,
            'Safety_Stock': transfer['Safety_Stock'],
            'MOQ': transfer['MOQ'],
            'Notes': f"{transfer['Type']} -> {receive['Type']}"
        }
    
    def calculate_statistics(self, suggestions):
        """計算統計分析"""
        if not suggestions:
            return {}
        
        df_suggestions = pd.DataFrame(suggestions)
        
        # 基本KPI
        stats = {
            'total_suggestions': len(df_suggestions),
            'total_qty': df_suggestions['Transfer_Qty'].sum(),
            'total_articles': df_suggestions['Article'].nunique(),
            'total_oms': df_suggestions['OM'].nunique(),
        }
        
        # 按產品統計
        article_stats = df_suggestions.groupby('Article').agg({
            'Transfer_Qty': 'sum',
            'Article': 'count',
            'OM': 'nunique'
        }).rename(columns={'Article': 'Count', 'OM': 'OM_Count'})
        
        # 按OM統計
        om_stats = df_suggestions.groupby('OM').agg({
            'Transfer_Qty': 'sum',
            'OM': 'count',
            'Article': 'nunique'
        }).rename(columns={'OM': 'Count', 'Article': 'Article_Count'})
        
        # 轉出類型分佈
        transfer_type_stats = df_suggestions.groupby('Transfer_Type').agg({
            'Transfer_Qty': 'sum',
            'Transfer_Type': 'count'
        }).rename(columns={'Transfer_Type': 'Count'})
        
        # 接收類型分佈
        receive_type_stats = df_suggestions.groupby('Receive_Type').agg({
            'Transfer_Qty': 'sum',
            'Receive_Type': 'count'
        }).rename(columns={'Receive_Type': 'Count'})
        
        stats.update({
            'article_stats': article_stats,
            'om_stats': om_stats,
            'transfer_type_stats': transfer_type_stats,
            'receive_type_stats': receive_type_stats
        })
        
        return stats
    
    def generate_recommendations(self, mode="A"):
        """生成調貨建議"""
        if self.df is None:
            return False, "請先載入數據"
        
        try:
            self.mode = mode
            
            # 識別候選
            transfer_candidates = self.identify_transfer_candidates(mode)
            receive_candidates = self.identify_receive_candidates()
            
            # 解決同店舖同SKU衝突 - v1.72 新增
            transfer_candidates, receive_candidates = self.resolve_same_store_conflicts(transfer_candidates, receive_candidates)
            
            # 匹配建議
            suggestions = self.match_transfer_suggestions(transfer_candidates, receive_candidates)
            
            # 計算統計
            statistics = self.calculate_statistics(suggestions)
            
            self.transfer_suggestions = suggestions
            self.statistics = statistics
            
            return True, f"成功生成 {len(suggestions)} 條調貨建議"
            
        except Exception as e:
            return False, f"生成建議失敗: {str(e)}"
    
    def create_visualization(self):
        """創建視覺化圖表"""
        if not self.transfer_suggestions:
            return None
        
        df_suggestions = pd.DataFrame(self.transfer_suggestions)
        
        # 按OM統計數據
        om_transfer_stats = df_suggestions.groupby(['OM', 'Transfer_Type'])['Transfer_Qty'].sum().unstack(fill_value=0)
        om_receive_stats = df_suggestions.groupby(['OM', 'Receive_Type'])['Transfer_Qty'].sum().unstack(fill_value=0)
        
        # 合併統計數據並重命名為英文
        om_stats = pd.concat([om_transfer_stats, om_receive_stats], axis=1, sort=False).fillna(0)
        
        # 重命名列為英文
        column_mapping = {
            'ND轉出': 'ND Transfer',
            'RF過剩轉出': 'RF Excess Transfer', 
            'RF加強轉出': 'RF Enhanced Transfer',
            '緊急缺貨補貨': 'Emergency Restock',
            'SasaNet調撥接收': 'SasaNet Transfer Receive',
            '潛在缺貨補貨': 'Potential Restock'
        }
        
        # 重命名存在的列
        om_stats.columns = [column_mapping.get(col, col) for col in om_stats.columns]
        
        # 創建圖表
        fig, ax = plt.subplots(figsize=(14, 8))
        
        # 設置條形圖位置
        x = np.arange(len(om_stats.index))
        width = 0.2
        
        # 定義顏色
        colors = ['#ff9999', '#66b3ff', '#99ff99', '#ffcc99', '#ffb3e6']
        
        # 繪製條形圖
        bars = []
        labels = []
        positions = []
        
        bar_position = 0
        for i, col in enumerate(om_stats.columns):
            if om_stats[col].sum() > 0:  # 只顯示有數據的類型
                bars.append(ax.bar(x + bar_position * width, om_stats[col], width, 
                                 label=col, color=colors[i % len(colors)]))
                labels.append(col)
                positions.append(bar_position)
                bar_position += 1
        
        # 設置圖表 - 全英文標籤
        ax.set_xlabel('OM Units', fontsize=12)
        ax.set_ylabel('Transfer Quantity', fontsize=12)
        
        if self.mode == "A":
            ax.set_title('OM Transfer vs Receive Analysis (Conservative Mode)', fontsize=14, fontweight='bold')
        else:
            ax.set_title('OM Transfer vs Receive Analysis (Enhanced Mode)', fontsize=14, fontweight='bold')
        
        ax.set_xticks(x + width * (len(positions) - 1) / 2)
        ax.set_xticklabels(om_stats.index, rotation=45, ha='right')
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1))
        ax.grid(axis='y', alpha=0.3)
        
        # 添加數值標籤
        for bar_group in bars:
            for bar in bar_group:
                height = bar.get_height()
                if height > 0:
                    ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                           f'{int(height)}', ha='center', va='bottom', fontsize=9)
        
        plt.tight_layout()
        return fig
    
    def export_to_excel(self):
        """匯出到Excel"""
        if not self.transfer_suggestions:
            return None, "沒有可匯出的數據"
        
        try:
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # 工作表1: 調貨建議
                df_suggestions = pd.DataFrame(self.transfer_suggestions)
                
                # 合併產品描述
                df_suggestions = df_suggestions.merge(
                    self.df[['Article', 'Article Description']].drop_duplicates(),
                    on='Article', how='left'
                )
                
                # 重新排列欄位
                export_columns = [
                    'Article', 'Article Description', 'OM', 'Transfer_Site', 
                    'Receive_Site', 'Transfer_Qty', 'Original_Stock', 
                    'After_Transfer_Stock', 'Safety_Stock', 'MOQ', 'Notes'
                ]
                
                df_export = df_suggestions[export_columns].rename(columns={
                    'Article': 'Article',
                    'Article Description': 'Product Desc',
                    'OM': 'OM',
                    'Transfer_Site': 'Transfer Site',
                    'Receive_Site': 'Receive Site',
                    'Transfer_Qty': 'Transfer Qty',
                    'Original_Stock': 'Original Stock',
                    'After_Transfer_Stock': 'After Transfer Stock',
                    'Safety_Stock': 'Safety Stock',
                    'MOQ': 'MOQ',
                    'Notes': 'Notes'
                })
                
                df_export.to_excel(writer, sheet_name='調貨建議', index=False)
                
                # 工作表2: 統計摘要
                stats_sheet = writer.book.create_sheet('統計摘要')
                row = 1
                
                # KPI概覽
                stats_sheet.cell(row=row, column=1, value="KPI概覽").font = Font(bold=True, size=14)
                row += 2
                
                kpi_data = [
                    ['總建議數', self.statistics['total_suggestions']],
                    ['總件數', self.statistics['total_qty']],
                    ['涉及產品數', self.statistics['total_articles']],
                    ['涉及OM數', self.statistics['total_oms']]
                ]
                
                for item in kpi_data:
                    stats_sheet.cell(row=row, column=1, value=item[0])
                    stats_sheet.cell(row=row, column=2, value=item[1])
                    row += 1
                
                row += 3
                
                # 其他統計表格
                stat_tables = [
                    ('按Article統計', self.statistics['article_stats']),
                    ('按OM統計', self.statistics['om_stats']),
                    ('轉出類型分佈', self.statistics['transfer_type_stats']),
                    ('接收類型分佈', self.statistics['receive_type_stats'])
                ]
                
                for title, df_stat in stat_tables:
                    stats_sheet.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
                    row += 2
                    
                    # 寫入表格標題
                    for col, header in enumerate(df_stat.columns):
                        stats_sheet.cell(row=row, column=col+2, value=header)
                    row += 1
                    
                    # 寫入數據
                    for idx, (index_val, series) in enumerate(df_stat.iterrows()):
                        stats_sheet.cell(row=row, column=1, value=index_val)
                        for col, val in enumerate(series):
                            stats_sheet.cell(row=row, column=col+2, value=val)
                        row += 1
                    
                    row += 3
            
            # 生成文件名
            date_str = datetime.now().strftime("%Y%m%d")
            filename = f"調貨建議_{date_str}.xlsx"
            
            output.seek(0)
            return output.getvalue(), filename
            
        except Exception as e:
            return None, f"匯出失敗: {str(e)}"

def main():
    """主應用程序"""
    
    # 頁面標題
    st.markdown('<h1 class="main-title">📦 調貨建議生成系統</h1>', unsafe_allow_html=True)
    st.markdown("---")
    
    # 初始化系統
    if 'system' not in st.session_state:
        st.session_state.system = TransferRecommendationSystem()
    
    system = st.session_state.system
    
    # 側邊欄
    with st.sidebar:
        st.header("系統資訊")
        st.info("""
**版本：v1.7a**  
**開發者: Ricky**

**核心功能：**  
- ✅ ND/RF類型智慧識別
- ✅ 優先順序調貨匹配
- ✅ RF過剩/加強轉出限制
- ✅ 統計分析和圖表
- ✅ Excel格式匯出
        """)
        
        st.markdown("---")
        
        # 轉貨模式選擇
        st.subheader("🎯 轉貨策略選擇")
        mode = st.radio(
            "選擇轉貨模式：",
            ["A: 保守轉貨", "B: 加強轉貨"],
            help="A模式：RF類型20%限制；B模式：RF類型50%限制，基於MOQ+1件"
        )
        
        transfer_mode = "A" if "A:" in mode else "B"
    
    # 主內容區域
    
    # 1. 資料上傳區塊
    st.markdown('<div class="section-header"><h2>📁 資料上傳</h2></div>', unsafe_allow_html=True)
    
    uploaded_file = st.file_uploader(
        "選擇Excel文件",
        type=['xlsx', 'xls'],
        help="請確保文件包含所有必需欄位"
    )
    
    if uploaded_file is not None:
        with st.spinner("正在載入資料..."):
            success, message = system.load_and_preprocess_data(uploaded_file)
            
        if success:
            st.success(message)
            
            # 2. 資料預覽區塊
            st.markdown('<div class="section-header"><h2>👀 資料預覽</h2></div>', unsafe_allow_html=True)
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("總記錄數", len(system.df))
            with col2:
                st.metric("產品數量", system.df['Article'].nunique())
            with col3:
                st.metric("店鋪數量", system.df['Site'].nunique())
            with col4:
                st.metric("OM數量", system.df['OM'].nunique())
            
            # 預計統計資訊
            if hasattr(system, 'preliminary_stats') and system.preliminary_stats:
                st.subheader("📊 預計統計資訊")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**A模式 (保守轉貨):**")
                    conservative = system.preliminary_stats['conservative']
                    subcol1, subcol2, subcol3 = st.columns(3)
                    with subcol1:
                        st.metric("預計轉出", conservative['estimated_transfer'])
                    with subcol2:
                        st.metric("預計接收", conservative['estimated_receive'])
                    with subcol3:
                        st.metric("預計需求", conservative['estimated_demand'])
                
                with col2:
                    st.markdown("**B模式 (加強轉貨):**")
                    enhanced = system.preliminary_stats['enhanced']
                    subcol1, subcol2, subcol3 = st.columns(3)
                    with subcol1:
                        st.metric("預計轉出", enhanced['estimated_transfer'])
                    with subcol2:
                        st.metric("預計接收", enhanced['estimated_receive'])
                    with subcol3:
                        st.metric("預計需求", enhanced['estimated_demand'])
            
            # 顯示資料樣本
            with st.expander("查看資料樣本", expanded=False):
                # 創建用於顯示的數據副本，確保清理所有異常字符
                display_df = system.df.head(10).copy()
                
                # 更严格地清理所有字符串列的顯示內容
                for col in display_df.columns:
                    if display_df[col].dtype == 'object':
                        import re
                        display_df[col] = display_df[col].astype(str).apply(
                            lambda x: 'HIDDEN_DATA' if re.search(r'key.*v_right|[\u0000-\u001f\u007f-\u009f]', str(x))
                            else re.sub(r'[^\w\s\u4e00-\u9fff\-_()./]', '', str(x)).strip()
                        )
                
                # 同时清理列名
                cleaned_display_columns = []
                for col in display_df.columns:
                    if re.search(r'key.*v_right|[\u0000-\u001f\u007f-\u009f]', str(col)):
                        cleaned_col = f"Unknown_Column_{len(cleaned_display_columns)}"
                    else:
                        cleaned_col = re.sub(r'[^\w\s\u4e00-\u9fff\-_()./]', '', str(col)).strip()
                        if not cleaned_col:
                            cleaned_col = f"Column_{len(cleaned_display_columns)}"
                    cleaned_display_columns.append(cleaned_col)
                
                display_df.columns = cleaned_display_columns
                
                st.dataframe(display_df, use_container_width=True)
            
            # 3. 分析按鈕區塊
            st.markdown('<div class="section-header"><h2>🔍 調貨分析</h2></div>', unsafe_allow_html=True)
            
            if st.button("🚀 生成調貨建議", type="primary", use_container_width=True):
                with st.spinner(f"正在分析調貨建議 ({mode})..."):
                    success, message = system.generate_recommendations(transfer_mode)
                
                if success:
                    st.success(message)
                    
                    # 4. 結果展示區塊
                    if system.transfer_suggestions:
                        st.markdown('<div class="section-header"><h2>📊 分析結果</h2></div>', unsafe_allow_html=True)
                        
                        # KPI指標卡
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("總建議數", system.statistics['total_suggestions'])
                        with col2:
                            st.metric("總調貨件數", system.statistics['total_qty'])
                        with col3:
                            st.metric("涉及產品", system.statistics['total_articles'])
                        with col4:
                            st.metric("涉及OM", system.statistics['total_oms'])
                        
                        # 調貨建議表格
                        st.subheader("📋 調貨建議明細")
                        df_display = pd.DataFrame(system.transfer_suggestions)
                        st.dataframe(df_display, use_container_width=True)
                        
                        # 統計分析表格
                        st.subheader("📈 統計分析")
                        
                        tab1, tab2, tab3, tab4 = st.tabs(["按產品統計", "按OM統計", "轉出類型分佈", "接收類型分佈"])
                        
                        with tab1:
                            st.dataframe(system.statistics['article_stats'], use_container_width=True)
                        
                        with tab2:
                            st.dataframe(system.statistics['om_stats'], use_container_width=True)
                        
                        with tab3:
                            st.dataframe(system.statistics['transfer_type_stats'], use_container_width=True)
                        
                        with tab4:
                            st.dataframe(system.statistics['receive_type_stats'], use_container_width=True)
                        
                        # 視覺化圖表
                        st.subheader("📊 視覺化分析")
                        fig = system.create_visualization()
                        if fig:
                            st.pyplot(fig)
                        
                        # 5. 匯出區塊
                        st.markdown('<div class="section-header"><h2>💾 匯出結果</h2></div>', unsafe_allow_html=True)
                        
                        excel_data, filename = system.export_to_excel()
                        if excel_data:
                            st.download_button(
                                label="📥 下載Excel報告",
                                data=excel_data,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        else:
                            st.error("匯出失敗")
                
                else:
                    st.error(message)
        
        else:
            st.error(message)
    
    else:
        st.info("👆 請上傳Excel文件開始分析")
        
        # 顯示文件格式說明
        with st.expander("📋 文件格式要求", expanded=True):
            st.markdown("""
            **必需欄位：**
            - `Article` - 產品編號
            - `Article Description` - 產品描述
            - `RP Type` - 補貨類型 (ND/RF)
            - `Site` - 店鋪編號
            - `OM` - 營運管理單位
            - `MOQ` - 最低派貨數量
            - `SaSa Net Stock` - 現有庫存數量
            - `Pending Received` - 在途訂單數量
            - `Safety Stock` - 安全庫存數量
            - `Last Month Sold Qty` - 上月銷量
            - `MTD Sold Qty` - 本月至今銷量
            
            **文件格式：** Excel (.xlsx, .xls)
            """)

if __name__ == "__main__":
    main()
